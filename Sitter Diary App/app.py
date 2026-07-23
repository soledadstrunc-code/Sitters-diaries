import hashlib
import html
import re
from datetime import datetime
from pathlib import Path

import gspread
import openpyxl
import streamlit as st
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

st.set_page_config(page_title="Provider Lifecycle Study — Sitter Profiles", layout="wide")

SEGMENT_LABELS = {
    "M": "MID", "T": "TOP", "N": "NEW", "P": "TOP",
    "MID": "MID", "TOP": "TOP", "NEW": "NEW", "PRO": "TOP",
}
SEGMENT_DISPLAY = {"TOP": "Pro Provider", "MID": "Mid Provider", "NEW": "New Provider"}
# Internal segment code stays "TOP" everywhere (filtering, dict keys, etc.) — this is
# only the short label shown to the user, e.g. on badges and the segment filter.
SEGMENT_BADGE_LABEL = {"TOP": "PRO", "MID": "MID", "NEW": "NEW"}
BADGE_COLORS = {"TOP": "#1E7A34", "MID": "#2455C4", "NEW": "#6B3FBF"}
BADGE_BG = {"TOP": "#E9F6EC", "MID": "#EAF1FF", "NEW": "#F3EEFB"}

CORE_FIELDS = {"date", "user name", "segment", "week #"}

# ---------------- CODED INTERVIEWS (local _Coded.xlsx workbooks) ----------------
# Provider subfolders (e.g. "Alexandra - New Provider") live inside "Kick off
# interviews", a sibling of this app's own folder under the "User Diaries" project
# folder. Each one may contain "<Name>_Coded.xlsx", produced by the
# sitter-interview-transcript-coding workflow. This reads each workbook's "Coding"
# sheet (one row per category) directly off disk, and also uses the folder names
# themselves as the master participant roster (see load_user_roster below) — no
# Google Sheets/Drive round trip needed for either.
#
# NOTE: the coding taxonomy itself has evolved mid-study — older workbooks use a
# 9-category scheme (Background, Motivations, Before booking, ...), newer ones
# (Pro providers so far) use a different 12-category scheme (Context and provider
# model, Motivation, Client acquisition, ...) and add Snapshot/Atomic sheets. Rather
# than hardcode one category list, load_coded_interview reads whatever category
# rows are actually present in each participant's own Coding sheet.

KICKOFF_DIR = Path(__file__).resolve().parent.parent / "Kick off interviews"


def _providers_dir():
    """Provider subfolders have sometimes lived directly under 'Kick off interviews'
    and sometimes under a 'Kick off interviews/Users' subfolder, depending on how the
    folder's been organized most recently. Checked fresh on every call (cheap — just a
    path check) so either layout works without code changes if it moves again."""
    nested = KICKOFF_DIR / "Users"
    return nested if nested.exists() else KICKOFF_DIR

# Known mismatches between the participant name used in the Diary Study Google Sheet
# and the folder/file name used under "Kick off interviews/Users" (folder names were
# set from transcript filenames/participant IDs assigned before the interviews). Add
# to this if more turn up — e.g. the "Ami" folder is actually about a participant
# named "Amy".
NAME_ALIASES = {"Amy": "Ami"}
_ALIAS_TO_CANONICAL = {alias.lower(): canonical for canonical, alias in NAME_ALIASES.items()}


def _segment_from_folder_name(folder_name):
    """Provider folder names look like '<Name> - New Provider' / '... - Pro Provider'
    / '... - Mid Provider' (with occasional 'Prodiver' typos and inconsistent spacing
    around the dash) — this pulls the segment word right after the dash."""
    parts = folder_name.split("-", 1)
    if len(parts) < 2:
        return "NEW"
    suffix_words = parts[1].strip().split()
    seg_word = suffix_words[0].upper() if suffix_words else ""
    return SEGMENT_LABELS.get(seg_word, "NEW")


def _find_coded_workbook_path(name):
    """Looks for <alias>_Coded.xlsx inside a per-provider subfolder of the local
    'Kick off interviews' folder (or its 'Users' subfolder, see _providers_dir).
    Folder names look like '<Name> - New Provider', so this matches on the leading
    name segment, case-insensitively, trying both the sheet's name and any known
    alias for it."""
    providers_dir = _providers_dir()
    if not providers_dir.exists():
        return None
    candidates = {name.strip().lower(), NAME_ALIASES.get(name.strip(), name.strip()).lower()}
    for child in providers_dir.iterdir():
        if not child.is_dir():
            continue
        folder_prefix = child.name.split("-")[0].strip().lower()
        if folder_prefix not in candidates:
            continue
        matches = list(child.glob("*_Coded.xlsx"))
        if matches:
            return matches[0]
    return None


@st.cache_data(ttl=300, show_spinner=False)
def load_snapshot(name):
    """Reads a participant's <Name>_Coded.xlsx 'Snapshot' sheet: a Field / Value /
    Supporting quote table capturing their segment, experience on Rover, income
    role, client counts, service types, growth intention, and other context.
    Returns None if there's no coded workbook yet, or it has no Snapshot sheet."""
    path = _find_coded_workbook_path(name)
    if not path:
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Snapshot" not in wb.sheetnames:
        return None

    rows = []
    for row in wb["Snapshot"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        field = str(row[0]).strip()
        value = row[1] if len(row) > 1 and row[1] else ""
        quote = row[2] if len(row) > 2 and row[2] else ""
        rows.append((field, str(value), str(quote)))
    return rows or None


_QUOTE_RE = re.compile(r'"([^"]*)"')


def gray_quotes(text):
    """Wraps double-quoted substrings (verbatim participant quotes) in Streamlit's
    gray-colored markdown span, so quotes read visually distinct from the summary
    prose around them."""
    return _QUOTE_RE.sub(lambda m: f':gray["{m.group(1)}"]', text)


def split_into_lines(text):
    """Coded summaries are either a bullet list (points separated by '•', with
    embedded newlines that Markdown collapses instead of rendering as line breaks)
    or a single run-on paragraph. This splits on bullets when present, otherwise on
    sentence-ending periods, so each point/sentence gets its own line rather than
    running together in one block."""
    if "•" in text:
        parts = [p.strip() for p in text.split("•") if p.strip()]
    else:
        parts = [s.strip() for s in re.split(r'(?<=\.)\s+(?=[A-Z"“])', text) if s.strip()]
    return parts or [text.strip()]


# Streamlit's default body text renders around 16px (~12pt); titles/labels below
# are sized a couple points bigger so they stand out from the surrounding text.
TITLE_FONT_SIZE = "1.15rem"

# One relevant emoji per Background category, covering both coding taxonomies used
# across the study (see the taxonomy note above load_coded_interview). Matched
# case-insensitively; anything unrecognized (a future taxonomy revision, etc.)
# falls back to a generic pin so the UI never breaks.
CATEGORY_EMOJIS = {
    # Older 9-category scheme
    "background": "👤",
    "motivations": "❤️",
    "before booking": "🔍",
    "booking process": "📅",
    "during the sitting": "🐾",
    "after the sitting": "✅",
    "pain points/frictions": "⚠️",
    "main needs or wishes": "🙏",
    "future: expectations": "🔮",
    # Newer 12-category scheme (Pro providers)
    "context and provider model": "👤",
    "motivation": "❤️",
    "client acquisition": "🔍",
    "evaluation and meet-and-greet": "🤝",
    "booking and availability": "📅",
    "sitting preparation": "🧳",
    "service delivery": "🐾",
    "client communication": "💬",
    "completion and feedback": "⭐",
    "payments and earnings": "💰",
    "cross-platform behavior": "🔗",
    "growth and future plans": "📈",
}
DEFAULT_CATEGORY_EMOJI = "📌"


def category_emoji(category):
    return CATEGORY_EMOJIS.get(category.strip().lower(), DEFAULT_CATEGORY_EMOJI)


def render_title(text):
    """Renders a standalone section/category heading at TITLE_FONT_SIZE."""
    st.markdown(
        f'<div style="font-size:{TITLE_FONT_SIZE}; font-weight:600; margin:10px 0 2px;">{text}</div>',
        unsafe_allow_html=True,
    )


def render_entries_table(rows, columns):
    """Renders a plain HTML table with fixed column widths, instead of st.table's
    auto-sizing (which was squeezing the Date column so narrow that dates wrapped
    across multiple lines). Short columns (like Date) stay on one line; the last
    column gets the remaining width and wraps normally so long Description text
    is fully readable without being cut off or forcing odd row heights elsewhere.

    `columns` is a list of (header, key, width_css) tuples; width_css is any valid
    CSS width value (e.g. "100px"), or None for the flexible last column."""
    thead = "".join(
        f'<th style="width:{w};text-align:left;padding:6px 10px;'
        f'border-bottom:1px solid #DADDE3;white-space:nowrap;">{html.escape(str(label))}</th>'
        if w else
        f'<th style="text-align:left;padding:6px 10px;border-bottom:1px solid #DADDE3;">{html.escape(str(label))}</th>'
        for label, _, w in columns
    )
    body_rows = []
    for row in rows:
        cells = []
        for label, key, w in columns:
            value = html.escape(str(row.get(key, ""))).replace("\n", "<br>")
            nowrap = "white-space:nowrap;" if w else "word-wrap:break-word;white-space:pre-wrap;"
            cells.append(
                f'<td style="padding:6px 10px;border-bottom:1px solid #F1F3F6;'
                f'vertical-align:top;{nowrap}">{value}</td>'
            )
        body_rows.append(f"<tr>{''.join(cells)}</tr>")

    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;font-size:13px;">'
        f'<thead><tr>{thead}</tr></thead><tbody>{"".join(body_rows)}</tbody></table>',
        unsafe_allow_html=True,
    )


@st.cache_data(ttl=300, show_spinner=False)
def load_coded_interview(name):
    """Reads a participant's <Name>_Coded.xlsx 'Coding' sheet from the local 'Kick
    off interviews' folder: one row per category, in whatever category scheme that
    particular workbook uses (see the taxonomy note above — this does not assume a
    fixed list of category names, since it varies by when the transcript was coded).
    Returns None if no coded workbook exists yet for this provider, so callers can
    fall back to a pending-state message."""
    path = _find_coded_workbook_path(name)
    if not path:
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Coding" not in wb.sheetnames:
        return None

    date_coded = ""
    if "Metadata" in wb.sheetnames:
        for row in wb["Metadata"].iter_rows(values_only=True):
            if row and row[0] and str(row[0]).strip().lower() == "date coded":
                date_coded = row[1] if len(row) > 1 and row[1] else ""
                break

    categories = []
    for row in wb["Coding"].iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            summary = row[1] if len(row) > 1 and row[1] else "Not discussed in this interview"
            categories.append((str(row[0]).strip(), str(summary)))

    if not categories:
        return None
    return {"dateCoded": str(date_coded) if date_coded else "", "categories": categories}


# ---------------- JOURNEY MAP (Pro_Segment_Patterns.xlsx) ----------------
@st.cache_data(ttl=300, show_spinner=False)
def load_journey_map():
    """Reads the 'Journey Map (Pro)' sheet from Pro_Segment_Patterns.xlsx in the
    local 'Kick off interviews' folder. That sheet stacks three tables in one grid:
    (1) the 10 sequential workflow stages with a description and an off-platform
    leakage note each; (2) every friction/need/workaround theme tagged in each
    stage, with a live distinct-provider count and a representative quote; (3)
    cross-cutting themes that show up as a friction/need/workaround in more than
    one stage. Columns I onward are a hidden helper block mirroring the raw atomic
    data for the sheet's own formulas — not used here. Returns None if the
    workbook or sheet isn't there yet."""
    path = KICKOFF_DIR / "Pro_Segment_Patterns.xlsx"
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Journey Map (Pro)" not in wb.sheetnames:
        return None

    rows = list(wb["Journey Map (Pro)"].iter_rows(values_only=True))

    # Section 1: stage backbone — rows right after the header, until a blank Stage.
    stages = []
    i = 1
    while i < len(rows) and rows[i][0]:
        stage, covers, leakage = rows[i][0], rows[i][1], rows[i][2]
        stages.append({"stage": str(stage), "covers": covers or "", "leakage": leakage or "", "items": []})
        i += 1
    stage_lookup = {s["stage"]: s for s in stages}

    # Section 2: "Frictions, Needs & Workarounds by Stage" — find its column-header
    # row (Stage/Theme/Type/...), then read data rows until the next blank Stage.
    j = i
    while j < len(rows) and rows[j][0] != "Stage":
        j += 1
    j += 1
    while j < len(rows) and rows[j][0]:
        r = rows[j]
        stage_name = str(r[0])
        if stage_name in stage_lookup:
            stage_lookup[stage_name]["items"].append({
                "theme": r[1], "type": r[2], "distinct_providers": r[3] or 0,
                "cross_cutting": r[4], "providers": r[5],
                "observation": r[6] or "", "quote": r[7] or "",
            })
        j += 1

    # Section 3: "Cross-Cutting Frictions, Needs & Workarounds" — find its
    # column-header row (Theme/Theme definition/...), then read until blank Theme.
    k = j
    while k < len(rows) and rows[k][0] != "Theme":
        k += 1
    k += 1
    cross_cutting = []
    while k < len(rows) and rows[k][0]:
        r = rows[k]
        cross_cutting.append({
            "theme": r[0], "definition": r[1] or "", "types": r[2] or "",
            "stage_count": r[3] or 0, "stages_detail": r[4] or "", "total_providers": r[5] or 0,
        })
        k += 1

    return {"stages": stages, "cross_cutting": cross_cutting}


# ---------------- AUTH ----------------
def check_login(email, password):
    auth_cfg = st.secrets.get("auth", {})
    allowed_domain = auth_cfg.get("allowed_domain", "").strip().lower()
    shared_hash = auth_cfg.get("shared_password_hash", "")

    email = email.strip().lower()
    if not allowed_domain or not shared_hash:
        return False
    if not email.endswith("@" + allowed_domain):
        return False
    return hashlib.sha256(password.encode()).hexdigest() == shared_hash


def render_login():
    st.title("Provider Lifecycle Study — Sitter Profiles")
    st.caption("Sign in with your Rover email to view the diary study dashboard.")
    with st.form("login_form"):
        email = st.text_input("Email", placeholder="you@rover.com")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in")
    if submitted:
        if check_login(email, password):
            st.session_state.authenticated = True
            st.session_state.user_email = email.strip().lower()
            st.rerun()
        else:
            st.error("Incorrect email or password, or your email isn't a Rover address.")


def render_logout_button():
    with st.sidebar:
        st.caption(f"Signed in as {st.session_state.get('user_email', '')}")
        if st.button("Log out"):
            for key in ("authenticated", "user_email", "view", "current_user"):
                st.session_state.pop(key, None)
            st.rerun()


# ---------------- GOOGLE SHEETS ----------------
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]


@st.cache_resource(show_spinner=False)
def get_gspread_client():
    info = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    return gspread.authorize(creds)


@st.cache_resource(show_spinner=False)
def get_drive_service():
    info = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


@st.cache_data(ttl=300, show_spinner=False)
def get_provider_description(name):
    """Looks for a Google Doc inside this provider's Drive subfolder (under the
    'Kick off interviews' folder) and returns its text as a description. Returns
    None if no folder / doc is found, or if the drive folder isn't configured."""
    drive_cfg = st.secrets.get("drive", {})
    root_folder_id = drive_cfg.get("provider_folder_id")
    if not root_folder_id:
        return None

    service = get_drive_service()

    safe_name = name.replace("'", "\\'")
    folder_query = (
        f"'{root_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder' "
        f"and name = '{safe_name}'"
    )
    folders = service.files().list(q=folder_query, fields="files(id, name)").execute().get("files", [])
    if not folders:
        return None
    provider_folder_id = folders[0]["id"]

    doc_query = (
        f"'{provider_folder_id}' in parents and mimeType = 'application/vnd.google-apps.document' "
        "and trashed = false"
    )
    docs = service.files().list(
        q=doc_query, fields="files(id, name, modifiedTime)", orderBy="modifiedTime desc"
    ).execute().get("files", [])
    if not docs:
        return None

    doc_id = docs[0]["id"]
    content = service.files().export(fileId=doc_id, mimeType="text/plain").execute()
    text = content.decode("utf-8") if isinstance(content, bytes) else content
    return text.strip() or None


DIARY_STUDY_XLSX = Path(__file__).resolve().parent.parent / "Rover_Diary_Study_Tool.xlsx"


def _find_header_row(all_values):
    """The Entries sheet may have a title/banner row above the real header row
    (e.g. 'ROVER DIARY STUDY — SYNTHESIS ANALYSIS TOOL'), so scan for the row that
    actually contains the expected column names instead of assuming row 0 is it."""
    for i, row_values in enumerate(all_values[:10]):
        normalized = {str(c).strip().lower() for c in row_values if c is not None}
        if "user name" in normalized and "date" in normalized:
            return i
    return 0


def _dedupe_headers(raw_headers):
    headers = []
    seen = {}
    for i, h in enumerate(raw_headers):
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    return headers


@st.cache_data(ttl=300, show_spinner=False)
def load_entries():
    """Loads the diary study Entries data. Prefers a local copy of
    Rover_Diary_Study_Tool.xlsx (sibling of this app's folder, under 'User
    Diaries') when present — no Google credentials needed — and only falls back
    to the live Google Sheet if that local file isn't there."""
    if DIARY_STUDY_XLSX.exists():
        return _load_entries_from_local_file(DIARY_STUDY_XLSX)
    return _load_entries_from_google_sheets()


def _load_entries_from_local_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    entries_ws = None
    for ws in wb.worksheets:
        if "entries" in ws.title.lower():
            entries_ws = ws
            break
    if entries_ws is None:
        raise RuntimeError(f"Could not find a sheet containing 'Entries' in {path.name}.")

    all_values = list(entries_ws.iter_rows(values_only=True))
    if not all_values:
        return []

    header_row_idx = _find_header_row(all_values)
    headers = _dedupe_headers(all_values[header_row_idx])

    records = []
    for row_values in all_values[header_row_idx + 1:]:
        if not any(str(cell).strip() for cell in row_values if cell is not None):
            continue
        padded = list(row_values) + [None] * (len(headers) - len(row_values))
        records.append(dict(zip(headers, padded)))

    cleaned = []
    for r in records:
        row = {}
        for k, v in r.items():
            if isinstance(v, datetime):
                value = v.strftime("%Y-%m-%d")
            elif isinstance(v, str):
                value = v.strip()
            elif v is None:
                value = ""
            else:
                value = v
            row[k.strip()] = value
        if row.get("User Name"):
            cleaned.append(row)
    return cleaned


def _load_entries_from_google_sheets():
    client = get_gspread_client()
    spreadsheet_id = st.secrets["sheet"]["spreadsheet_id"]
    sh = client.open_by_key(spreadsheet_id)

    entries_ws = None
    for ws in sh.worksheets():
        if "entries" in ws.title.lower():
            entries_ws = ws
            break
    if entries_ws is None:
        raise RuntimeError("Could not find a tab containing 'Entries' in the spreadsheet.")

    all_values = entries_ws.get_all_values()
    if not all_values:
        return []

    header_row_idx = _find_header_row(all_values)
    headers = _dedupe_headers(all_values[header_row_idx])

    records = []
    for row_values in all_values[header_row_idx + 1:]:
        if not any(cell.strip() for cell in row_values):
            continue
        padded = row_values + [""] * (len(headers) - len(row_values))
        records.append(dict(zip(headers, padded)))

    # Normalize keys for easier lookup while keeping original labels for display
    cleaned = []
    for r in records:
        row = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
        if row.get("User Name"):
            cleaned.append(row)
    return cleaned


@st.cache_data(ttl=300, show_spinner=False)
def load_user_roster():
    """Builds the master participant list straight from the provider subfolders
    inside 'Kick off interviews' (or its 'Users' subfolder, see _providers_dir) —
    one folder per provider, e.g. 'Alexandra - New Provider'. This is the full study
    roster, so every enrolled provider gets a dashboard card even before they have
    any rows in the Diary Study Google Sheet. Returns {} if that folder isn't there,
    so the caller falls back to building the list purely from Sheet entries."""
    providers_dir = _providers_dir()
    if not providers_dir.exists():
        return {}
    roster = {}
    for child in sorted(providers_dir.iterdir()):
        if not child.is_dir():
            continue
        name = child.name.split("-", 1)[0].strip()
        if not name:
            continue
        name = _ALIAS_TO_CANONICAL.get(name.lower(), name)  # e.g. folder "Ami" -> roster name "Amy"
        roster[name] = _segment_from_folder_name(child.name)
    return roster


def build_participants(entries):
    roster = load_user_roster()
    participants = {
        name: {"name": name, "segment": segment, "entries": [], "country": "", "age": ""}
        for name, segment in roster.items()
    }

    for row in entries:
        name = row.get("User Name", "").strip()
        if not name:
            continue
        seg_raw = str(row.get("Segment", "")).strip().upper()
        segment = SEGMENT_LABELS.get(seg_raw, seg_raw or "NEW")
        if name not in participants:
            participants[name] = {"name": name, "segment": segment, "entries": [], "country": "", "age": ""}
        participants[name]["entries"].append(row)
        # Country/Age aren't in the Entries sheet today; this picks them up
        # automatically if those columns are ever added, without erroring otherwise.
        if not participants[name]["country"] and row.get("Country"):
            participants[name]["country"] = row["Country"]
        if not participants[name]["age"] and row.get("Age"):
            participants[name]["age"] = row["Age"]

    for p in participants.values():
        dates = [r.get("Date") for r in p["entries"] if r.get("Date")]
        p["entry_count"] = len(p["entries"])
        p["first_date"] = min(dates) if dates else None
        p["last_date"] = max(dates) if dates else None
    return participants


def badge_html(segment):
    color = BADGE_COLORS.get(segment, "#4B5565")
    bg = BADGE_BG.get(segment, "#F1F3F6")
    label = SEGMENT_BADGE_LABEL.get(segment, segment)
    return f'<span style="font-size:11px;font-weight:700;padding:3px 8px;border-radius:6px;background:{bg};color:{color};">{label}</span>'


def classify_entry(row):
    """Buckets each Entries row into one of two sections shown on a provider's
    profile (reflection vs. regular diary activity). Week-0 kickoff rows are
    excluded entirely (they're superseded by the coded interview Snapshot/
    Background tabs), everything else is a real diary entry.

    NOTE: this used to also treat any row whose Question started with "Topic" as
    a Week-0-only background row, but ordinary ongoing diary entries are also
    phrased as topic questions — that was silently dropping real diary entries
    (e.g. Deborah's) into a bucket nothing displays. Only "Week #" == "0" is used
    to detect kickoff rows now."""
    question = str(row.get("Question", "")).lower()
    task_type = str(row.get("Task Type", "")).lower()
    experience = str(row.get("Experience Moment", "")).lower()
    week = str(row.get("Week #", "")).strip()

    if "reflect" in question or "reflect" in task_type or "reflect" in experience:
        return "reflection"
    if week == "0":
        return "background"
    return "activity"


# ---------------- NAVIGATION ----------------
def go_to_profile(name):
    st.session_state.view = "profile"
    st.session_state.current_user = name


def go_to_dash():
    st.session_state.view = "dash"
    st.session_state.current_user = None


# ---------------- DASHBOARD VIEW ----------------
def render_dashboard(participants):
    st.title("Provider Lifecycle Study — Sitter Profiles")

    tab_profiles, tab_patterns = st.tabs(["User Profiles", "User Patterns"])

    with tab_profiles:
        render_user_profiles_tab(participants)

    with tab_patterns:
        # Journey map hidden for now (render_user_patterns_tab still has the full
        # implementation below — just not called — so it's a one-line change to
        # bring back).
        st.info("Coming soon — this tab will surface cross-participant patterns and themes.")


def render_user_patterns_tab():
    try:
        journey = load_journey_map()
    except Exception:
        journey = None

    if not journey:
        st.info(
            "No journey map found yet. Once \"Pro_Segment_Patterns.xlsx\" exists in the "
            "\"Kick off interviews\" folder with a \"Journey Map (Pro)\" sheet, it will show here."
        )
        return

    st.caption(
        "Pro segment journey map — synthesized across 14 Pro providers' coded interviews "
        "(Pro_Segment_Patterns.xlsx)."
    )

    last_stage_index = len(journey["stages"]) - 1
    for i, stage in enumerate(journey["stages"]):
        render_title(f"{category_emoji(stage['stage'])} {stage['stage']}")
        if stage["covers"]:
            st.markdown(gray_quotes(stage["covers"]))

        items = sorted(stage["items"], key=lambda it: it["distinct_providers"], reverse=True)
        if items:
            st.markdown("**Top frictions, needs & workarounds**")
            for it in items:
                st.markdown(
                    f"- **{it['theme']}** ({it['type']}, {it['distinct_providers']} of 14 providers) — "
                    f"{gray_quotes(it['observation'])}"
                )

        leakage = stage["leakage"]
        if leakage and not leakage.lower().startswith("no off-platform-clients-leakage"):
            st.caption(f"Off-platform leakage: {gray_quotes(leakage)}")

        if i < last_stage_index:
            st.divider()

    if journey["cross_cutting"]:
        st.divider()
        render_title("🔗 Cross-cutting themes")
        st.caption("Themes tagged as a friction, need, or workaround in more than one stage.")
        last_cc_index = len(journey["cross_cutting"]) - 1
        for i, theme in enumerate(journey["cross_cutting"]):
            st.markdown(f"**{theme['theme']}** — {theme['definition']}")
            st.caption(
                f"{theme['total_providers']} of 14 providers · spans {theme['stage_count']} stages: "
                f"{theme['stages_detail']}"
            )
            if i < last_cc_index:
                st.divider()


def render_user_profiles_tab(participants):
    st.caption(
        "Data refreshes automatically from Rover_Diary_Study_Tool (fed from MyInsights). "
        "Select a provider to see their kickoff background, weekly reflections, and daily entries."
    )

    top_col1, top_col2 = st.columns([5, 1])
    with top_col1:
        st.write("")
    with top_col2:
        if st.button("Refresh data", use_container_width=True):
            load_entries.clear()
            load_user_roster.clear()
            load_coded_interview.clear()
            load_snapshot.clear()
            load_journey_map.clear()
            st.rerun()

    st.markdown("**Filter by segment**")
    seg_filter_labels = {"ALL": "All", "NEW": "New", "MID": "Mid", "TOP": "Pro"}
    seg_filter = st.radio(
        "Segment", ["ALL", "NEW", "MID", "TOP"], horizontal=True, label_visibility="collapsed",
        format_func=lambda code: seg_filter_labels.get(code, code),
    )

    filtered = [p for p in participants.values() if seg_filter == "ALL" or p["segment"] == seg_filter]
    filtered.sort(key=lambda p: p["name"])

    if not filtered:
        st.info("No providers match your filters yet.")
        return

    st.write("")
    cols = st.columns(4)
    for i, p in enumerate(filtered):
        reflection_count = sum(1 for r in p["entries"] if classify_entry(r) == "reflection")
        with cols[i % 4]:
            with st.container(border=True):
                st.markdown(
                    f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                    f'<b>{p["name"]}</b>{badge_html(p["segment"])}</div>',
                    unsafe_allow_html=True,
                )
                st.caption(f"{p['entry_count']} entries · {reflection_count} weekly reflections")
                st.button(
                    "View profile", key=f"btn_{p['name']}",
                    on_click=go_to_profile, args=(p["name"],), use_container_width=True,
                )

    st.divider()
    st.caption(
        f"{len(participants)} providers total, sourced live from the Entries tab of Rover_Diary_Study_Tool. "
        "Segment codes: N = new, M = mid, P = pro."
    )


# ---------------- PROFILE VIEW ----------------
def render_profile(participants):
    name = st.session_state.current_user
    p = participants.get(name)

    st.button("← All providers", on_click=go_to_dash)

    if not p:
        st.warning(f"No data found for {name}. It may have been removed from the sheet.")
        return

    st.caption(
        f"{p['entry_count']} entries · {p['first_date'] or '—'} to {p['last_date'] or '—'}"
    )

    header_col1, header_col2 = st.columns([5, 1])
    with header_col1:
        segment_label = SEGMENT_DISPLAY.get(p["segment"], p["segment"])
        st.header(f"{p['name']} - {segment_label}")
    with header_col2:
        st.markdown(badge_html(p["segment"]), unsafe_allow_html=True)

    try:
        description = get_provider_description(p["name"])
    except Exception:
        description = None
    if description:
        st.write(description)
    else:
        st.caption(
            "No description yet — add a Google Doc to this provider's folder inside "
            "\"Kick off interviews\" in Drive and it will show up here."
        )

    entries = p["entries"]
    reflection_entries = [r for r in entries if classify_entry(r) == "reflection"]
    activity_entries = [r for r in entries if classify_entry(r) == "activity"]

    try:
        snapshot = load_snapshot(p["name"])
    except Exception:
        snapshot = None

    try:
        coded = load_coded_interview(p["name"])
    except Exception:
        coded = None

    tab_snapshot, tab_background, tab_diary, tab_reflections = st.tabs(
        ["Snapshot", "Background", "Daily entries", "Weekly reflections"]
    )

    with tab_snapshot:
        if not snapshot:
            st.info(
                f"No coded interview found yet for {p['name']}. Once their transcript is coded, "
                "this tab will show their segment, experience on Rover, income role, client counts, "
                "service types, and growth intention — straight from the transcript."
            )
        else:
            fields = [f for f in snapshot if f[0].strip().lower() != "provider segment"]
            for field, value, quote in fields:
                render_title(field)
                for line in split_into_lines(str(value)):
                    st.markdown(gray_quotes(line))

    with tab_background:
        if not coded:
            st.info(
                f"No coded interview found yet for {p['name']}. Once their transcript is coded "
                "into a <Name>_Coded.xlsx workbook inside their subfolder of \"Kick off interviews\", "
                "this tab will show their background, motivations, and workflow across the booking "
                "journey (before booking, booking process, during, after), pain points, needs, and "
                "future expectations — straight from the transcript."
            )
        else:
            if coded["dateCoded"]:
                st.caption(f"Coded from interview transcript · {coded['dateCoded']}")
            last_index = len(coded["categories"]) - 1
            for i, (category, summary) in enumerate(coded["categories"]):
                render_title(f"{category_emoji(category)} {category}")
                for line in split_into_lines(summary):
                    st.markdown(gray_quotes(line))
                if i < last_index:
                    st.divider()

    with tab_diary:
        if not activity_entries:
            st.info("No daily entries logged yet for this provider.")
        else:
            st.caption(f"{len(activity_entries)} daily entries")
            render_entries_table(
                activity_entries,
                columns=[
                    ("Date", "Date", "100px"),
                    ("Question", "Question", "220px"),
                    ("Description", "Description", None),
                ],
            )

    with tab_reflections:
        st.caption(f"{len(reflection_entries)} weekly reflections")
        if not reflection_entries:
            st.info(
                "No weekly reflections logged yet for this provider. These will appear here once "
                "loaded into MyInsights and synced to the Entries tab."
            )
        else:
            for row in reflection_entries:
                with st.container(border=True):
                    label = row.get("Week #")
                    render_title(f"Week {label}" if label not in (None, "") else "Weekly reflection")
                    for field, value in row.items():
                        key = field.strip().lower()
                        if key in CORE_FIELDS or not value:
                            continue
                        st.caption(f"**{field}:** {value}")


# ---------------- ROUTER ----------------
def main():
    if not st.session_state.get("authenticated"):
        render_login()
        return

    render_logout_button()

    # No secrets check here: load_entries() reads Rover_Diary_Study_Tool.xlsx
    # locally when it's present and only needs gcp_service_account/sheet secrets
    # in the Google Sheets fallback path, so requiring them upfront would block
    # anyone running purely off the local file.
    try:
        entries = load_entries()
    except Exception as e:
        st.error(f"Could not load diary study data: {e}")
        return

    participants = build_participants(entries)

    if "view" not in st.session_state:
        st.session_state.view = "dash"
    if "current_user" not in st.session_state:
        st.session_state.current_user = None

    if st.session_state.view == "dash":
        render_dashboard(participants)
    else:
        render_profile(participants)


if __name__ == "__main__":
    main()
